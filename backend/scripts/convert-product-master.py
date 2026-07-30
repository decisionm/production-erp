"""Convert the factory product master workbook to the importer's JSON rows.

Cells are copied VERBATIM: "18/20" and "21.5 / 17.8" must survive whole so the
importer can split them into unresolved variants. Whitespace-only cells become
null, because only genuinely blank values are meant to be flagged.

Columns are named by their 1-based SHEET column so the mapping can be read
straight off the workbook, rather than as offsets into a row slice.
"""
import json
import sys

import openpyxl

SHEET = 'PRODUCT DETAILS - SOFTWARE  (2'
FIRST_DATA_ROW, LAST_DATA_ROW = 10, 112

# 1-based sheet columns. Header rows 7-8 read:
#   B SL.NO. | C PRODUCT | D NO. OF CAVITY | E WT. | F CYCLE TIME
#   G BOTL/POUCH | H BOT/BOX | I POUCH/BOX DETAILS
#   J BOTL/TRAY | K BOT/BOX | L TRAY/BOX DETAILS
#   M CARTON | N TRAY | O POUCH
COLUMNS = {
    'sl_no': 2,
    'product': 3,
    'cavities': 4,
    'unit_weight_grams': 5,
    'cycle_time': 6,
    'nos_per_pouch': 7,
    'pouch_nos_per_box': 8,
    'pouches_per_box': 9,
    'nos_per_tray': 10,
    'tray_nos_per_box': 11,
    'trays_per_box': 12,
    'carton_spec': 13,
    'tray_spec': 14,
    'pouch_spec': 15,
}


def cell(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (int, float)):
        return str(value)
    return str(value).strip() or None


def main(xlsx_path: str, out_path: str) -> None:
    ws = openpyxl.load_workbook(xlsx_path, data_only=True)[SHEET]
    rows = []
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        row = {key: cell(ws.cell(row=r, column=col).value) for key, col in COLUMNS.items()}
        # The serial stays a number, as the committed fixture has it — it is a
        # row reference, never a value the importer splits or parses.
        raw_sl = ws.cell(row=r, column=COLUMNS['sl_no']).value
        row['sl_no'] = int(raw_sl) if isinstance(raw_sl, (int, float)) else raw_sl
        # A wholly empty trailing row is not a product.
        if row['product'] is None and row['sl_no'] is None:
            continue
        rows.append(row)

    with open(out_path, 'w') as fh:
        json.dump(rows, fh, indent=1)
    print(f'{len(rows)} rows -> {out_path}')


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
